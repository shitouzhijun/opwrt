#!/usr/bin/env python3
"""
专业记账模板生成器
创建包含多个工作表的专业记账Excel模板
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar

def create_professional_accounting_template():
    """创建专业记账Excel模板"""
    
    # 创建新的工作簿
    wb = Workbook()
    
    # 删除默认工作表
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ==================== 1. 首页 - 仪表板 ====================
    dashboard = wb.create_sheet("仪表板")
    dashboard.sheet_view.showGridLines = False
    
    # 设置列宽
    for col in range(1, 13):
        dashboard.column_dimensions[get_column_letter(col)].width = 12
    
    # 标题
    dashboard.merge_cells('A1:L1')
    title_cell = dashboard['A1']
    title_cell.value = "个人/家庭财务管理系统"
    title_cell.font = Font(name='微软雅黑', size=20, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill('solid', start_color='2E75B6')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 当前日期
    current_date = datetime.now().strftime('%Y年%m月%d日')
    dashboard['A3'] = f"报表日期: {current_date}"
    dashboard['A3'].font = Font(name='微软雅黑', size=11, bold=True)
    
    # 关键指标区域
    metrics = [
        ("A5", "当月总收入", "=SUMIFS(流水账!F:F,流水账!E:E,\"收入\",流水账!B:B,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),流水账!B:B,\"<=\"&EOMONTH(TODAY(),0))"),
        ("A7", "当月总支出", "=SUMIFS(流水账!F:F,流水账!E:E,\"支出\",流水账!B:B,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),流水账!B:B,\"<=\"&EOMONTH(TODAY(),0))"),
        ("A9", "当月结余", "=A6-A8"),
        ("E5", "年度总收入", "=SUMIFS(流水账!F:F,流水账!E:E,\"收入\",流水账!B:B,\">=\"&DATE(YEAR(TODAY()),1,1),流水账!B:B,\"<=\"&DATE(YEAR(TODAY()),12,31))"),
        ("E7", "年度总支出", "=SUMIFS(流水账!F:F,流水账!E:E,\"支出\",流水账!B:B,\">=\"&DATE(YEAR(TODAY()),1,1),流水账!B:B,\"<=\"&DATE(YEAR(TODAY()),12,31))"),
        ("E9", "年度结余", "=E6-E8"),
        ("I5", "当前总资产", "=SUM(资产负债表!C:C)"),
        ("I7", "当前总负债", "=SUM(资产负债表!F:F)"),
        ("I9", "净资产", "=I6-I8")
    ]
    
    for cell, label, formula in metrics:
        # 标签
        label_cell = dashboard[cell]
        label_cell.value = label
        label_cell.font = Font(name='微软雅黑', size=11, bold=True)
        label_cell.fill = PatternFill('solid', start_color='F2F2F2')
        
        # 数值
        value_cell = dashboard[cell.replace('A', 'B').replace('E', 'F').replace('I', 'J')]
        value_cell.value = formula
        value_cell.number_format = '#,##0.00'
        value_cell.font = Font(name='Arial', size=11, bold=True)
        value_cell.fill = PatternFill('solid', start_color='FFFFFF')
        
        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        label_cell.border = thin_border
        value_cell.border = thin_border
    
    # 月度支出分类饼图数据区域
    dashboard['A12'] = "月度支出分类"
    dashboard['A12'].font = Font(name='微软雅黑', size=12, bold=True, color='2E75B6')
    
    categories = ["餐饮", "交通", "购物", "娱乐", "住房", "医疗", "教育", "其他"]
    for i, category in enumerate(categories):
        row = 13 + i
        dashboard[f'A{row}'] = category
        dashboard[f'A{row}'].font = Font(name='微软雅黑', size=10)
        dashboard[f'B{row}'] = f'=SUMIFS(流水账!F:F,流水账!E:E,"支出",流水账!C:C,A{row},流水账!B:B,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),流水账!B:B,"<="&EOMONTH(TODAY(),0))'
        dashboard[f'B{row}'].number_format = '#,##0.00'
    
    # 使用说明
    dashboard['E12'] = "使用说明"
    dashboard['E12'].font = Font(name='微软雅黑', size=12, bold=True, color='2E75B6')
    
    instructions = [
        ("E13", "1. 在'流水账'表中记录每日收支"),
        ("E14", "2. 在'预算管理'表中设置预算"),
        ("E15", "3. 在'分类设置'中自定义分类"),
        ("E16", "4. 在'资产负债表'中记录资产"),
        ("E17", "5. 所有图表自动更新"),
        ("E18", "6. 使用数据验证确保数据准确")
    ]
    
    for cell, text in instructions:
        dashboard[cell] = text
        dashboard[cell].font = Font(name='微软雅黑', size=10)
    
    # ==================== 2. 流水账工作表 ====================
    ledger = wb.create_sheet("流水账")
    
    # 设置列宽
    column_widths = {'A': 12, 'B': 15, 'C': 20, 'D': 40, 'E': 10, 'F': 15, 'G': 20}
    for col, width in column_widths.items():
        ledger.column_dimensions[col].width = width
    
    # 标题行
    headers = ["序号", "日期", "分类", "项目描述", "类型", "金额", "支付方式"]
    for col, header in enumerate(headers, 1):
        cell = ledger.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 示例数据
    sample_data = [
        [1, datetime(2026, 3, 1), "餐饮", "午餐", "支出", 45.00, "微信支付"],
        [2, datetime(2026, 3, 1), "交通", "地铁卡充值", "支出", 100.00, "支付宝"],
        [3, datetime(2026, 3, 2), "工资", "3月工资", "收入", 15000.00, "银行转账"],
        [4, datetime(2026, 3, 2), "购物", "购买书籍", "支出", 89.00, "信用卡"],
        [5, datetime(2026, 3, 3), "娱乐", "电影票", "支出", 120.00, "微信支付"],
        [6, datetime(2026, 3, 4), "餐饮", "晚餐", "支出", 68.00, "支付宝"],
        [7, datetime(2026, 3, 5), "其他收入", "兼职收入", "收入", 800.00, "现金"]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ledger.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # 设置格式
            if col_idx == 2:  # 日期列
                cell.number_format = 'YYYY-MM-DD'
            elif col_idx == 6:  # 金额列
                cell.number_format = '#,##0.00'
                if row_data[4] == "支出":
                    cell.font = Font(color='FF0000')  # 支出用红色
                else:
                    cell.font = Font(color='00B050')  # 收入用绿色
    
    # 添加公式行
    total_row = len(sample_data) + 2
    ledger.cell(row=total_row, column=5, value="总收入:")
    ledger.cell(row=total_row, column=5).font = Font(bold=True)
    ledger.cell(row=total_row, column=6, value='=SUMIF(F:F,"收入",G:G)')
    ledger.cell(row=total_row, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row, column=6).font = Font(bold=True, color='00B050')
    
    ledger.cell(row=total_row+1, column=5, value="总支出:")
    ledger.cell(row=total_row+1, column=5).font = Font(bold=True)
    ledger.cell(row=total_row+1, column=6, value='=SUMIF(F:F,"支出",G:G)')
    ledger.cell(row=total_row+1, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row+1, column=6).font = Font(bold=True, color='FF0000')
    
    ledger.cell(row=total_row+2, column=5, value="结余:")
    ledger.cell(row=total_row+2, column=5).font = Font(bold=True)
    ledger.cell(row=total_row+2, column=6, value=f'={get_column_letter(6)}{total_row}-{get_column_letter(6)}{total_row+1}')
    ledger.cell(row=total_row+2, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row+2, column=6).font = Font(bold=True)
    
    # ==================== 3. 预算管理工作表 ====================
    budget = wb.create_sheet("预算管理")
    
    # 设置列宽
    budget.column_dimensions['A'].width = 25
    budget.column_dimensions['B'].width = 15
    budget.column_dimensions['C'].width = 15
    budget.column_dimensions['D'].width = 15
    budget.column_dimensions['E'].width = 15
    budget.column_dimensions['F'].width = 20
    
    # 标题
    budget['A1'] = "月度预算管理"
    budget['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')
    budget.merge_cells('A1:F1')
    
    # 表头
    budget_headers = ["预算分类", "月度预算", "实际支出", "差额", "完成率", "状态"]
    for col, header in enumerate(budget_headers, 1):
        cell = budget.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='5B9BD5')
        cell.alignment = Alignment(horizontal='center')
    
    # 预算分类数据
    budget_categories = [
        ["餐饮", 1500],
        ["交通", 800],
        ["购物", 2000],
        ["娱乐", 500],
        ["住房", 3000],
        ["医疗", 300],
        ["教育", 1000],
        ["其他", 500]
    ]
    
    for row_idx, (category, amount) in enumerate(budget_categories, 4):
        # 分类
        budget.cell(row=row_idx, column=1, value=category)
        
        # 月度预算
        budget.cell(row=row_idx, column=2, value=amount)
        budget.cell(row=row_idx, column=2).number_format = '#,##0.00'
        
        # 实际支出公式
        actual_formula = f'=SUMIFS(流水账!F:F,流水账!C:C,A{row_idx},流水账!E:E,"支出",流水账!B:B,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),流水账!B:B,"<="&EOMONTH(TODAY(),0))'
        budget.cell(row=row_idx, column=3, value=actual_formula)
        budget.cell(row=row_idx, column=3).number_format = '#,##0.00'
        
        # 差额公式
        diff_formula = f'=B{row_idx}-C{row_idx}'
        budget.cell(row=row_idx, column=4, value=diff_formula)
        budget.cell(row=row_idx, column=4).number_format = '#,##0.00'
        
        # 完成率公式
        completion_formula = f'=IF(B{row_idx}>0,C{row_idx}/B{row_idx},0)'
        budget.cell(row=row_idx, column=5, value=completion_formula)
        budget.cell(row=row_idx, column=5).number_format = '0.0%'
        
        # 状态公式
        status_formula = f'=IF(D{row_idx}>=0,"在预算内",IF(D{row_idx}>=-B{row_idx}*0.1,"接近超支","已超支"))'
        budget.cell(row=row_idx, column=6, value=status_formula)
    
    # 总计行
    total_row = len(budget_categories) + 4
    budget.cell(row=total_row, column=1, value="总计")
    budget.cell(row=total_row, column=1).font = Font(bold=True)
    
    for col in [2, 3, 4]:
        budget.cell(row=total_row, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})')
        budget.cell(row=total_row, column=col).number_format = '#,##0.00'
        budget.cell(row=total_row, column=col).font = Font(bold=True)
    
    # ==================== 4. 分类设置工作表 ====================
    categories = wb.create_sheet("分类设置")
    
    categories['A1'] = "收支分类设置"
    categories['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')
    categories.merge_cells('A1:C1')
    
    # 收入分类
    categories['A3'] = "收入分类"
    categories['A3'].font = Font(name='微软雅黑', size=12, bold=True)
    
    income_cats = ["工资", "奖金", "投资收入", "兼职收入", "其他收入", "退款", "报销"]
    for i, cat in enumerate(income_cats, 4):
        categories.cell(row=i, column=1, value=cat)
    
    # 支出分类
    categories['B3'] = "支出分类"
    categories['B3'].font = Font(name='微软雅黑', size=12, bold=True)
    
    expense_cats = ["餐饮", "交通", "购物", "娱乐", "住房", "医疗", "教育", "通讯", "保险", "其他"]
    for i, cat in enumerate(expense_cats, 4):
        categories.cell(row=i, column=2, value=cat)
    
    # 支付方式
    categories['C3'] = "支付方式"
    categories['C3'].font = Font(name='微软雅黑', size=12, bold=True)
    
    payment_methods = ["现金", "微信支付", "支付宝", "信用卡", "银行转账", "其他"]
    for i, method in enumerate(payment_methods, 4):
        categories.cell(row=i, column=3, value=method)
    
    # ==================== 5. 资产负债表 ====================
    balance_sheet = wb.create_sheet("资产负债表")
    
    balance_sheet['A1'] = "个人资产负债表"
    balance_sheet['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')
    balance_sheet.merge_cells('A1:F1')
    
    # 资产表头
    balance_sheet['A3'] = "资产"
    balance_sheet['A3'].font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    balance_sheet['A3'].fill = PatternFill('solid', start_color='70AD47')
    balance_sheet.merge_cells('A3:C3')
    
    asset_headers = ["资产项目", "金额", "备注"]
    for col, header in enumerate(asset_headers, 1):
        cell = balance_sheet.cell(row=