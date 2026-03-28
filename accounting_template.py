#!/usr/bin/env python3
"""
简化版专业记账模板生成器
使用纯openpyxl创建Excel记账模板
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_simple_accounting_template():
    """创建简化版专业记账Excel模板"""
    
    print("📊 开始创建专业记账模板...")
    
    # 创建新的工作簿
    wb = Workbook()
    
    # ==================== 1. 首页 - 仪表板 ====================
    print("  创建仪表板...")
    dashboard = wb.active
    dashboard.title = "仪表板"
    dashboard.sheet_view.showGridLines = False
    
    # 设置列宽
    for col in range(1, 13):
        dashboard.column_dimensions[get_column_letter(col)].width = 12
    
    # 标题
    dashboard.merge_cells('A1:L1')
    title_cell = dashboard['A1']
    title_cell.value = "个人/家庭财务管理系统"
    title_cell.font = Font(name='Arial', size=20, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill('solid', start_color='2E75B6')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 当前日期
    current_date = datetime.now().strftime('%Y年%m月%d日')
    dashboard['A3'] = f"报表日期: {current_date}"
    dashboard['A3'].font = Font(name='Arial', size=11, bold=True)
    
    # 关键指标
    metrics = [
        ("A5", "当月总收入", "收入总和"),
        ("A7", "当月总支出", "支出总和"),
        ("A9", "当月结余", "收入-支出"),
        ("E5", "年度总收入", "年度收入"),
        ("E7", "年度总支出", "年度支出"),
        ("E9", "年度结余", "年度结余"),
        ("I5", "当前总资产", "资产合计"),
        ("I7", "当前总负债", "负债合计"),
        ("I9", "净资产", "资产-负债")
    ]
    
    for cell, label, placeholder in metrics:
        # 标签
        label_cell = dashboard[cell]
        label_cell.value = label
        label_cell.font = Font(name='Arial', size=11, bold=True)
        label_cell.fill = PatternFill('solid', start_color='F2F2F2')
        
        # 数值占位符
        value_cell = dashboard[cell.replace('A', 'B').replace('E', 'F').replace('I', 'J')]
        value_cell.value = placeholder
        value_cell.number_format = '#,##0.00'
        value_cell.font = Font(name='Arial', size=11, bold=True)
        value_cell.fill = PatternFill('solid', start_color='FFFFFF')
    
    # 月度支出分类
    dashboard['A12'] = "月度支出分类"
    dashboard['A12'].font = Font(name='Arial', size=12, bold=True, color='2E75B6')
    
    categories = ["餐饮", "交通", "购物", "娱乐", "住房", "医疗", "教育", "其他"]
    for i, category in enumerate(categories):
        row = 13 + i
        dashboard[f'A{row}'] = category
        dashboard[f'A{row}'].font = Font(name='Arial', size=10)
        dashboard[f'B{row}'].value = 0
        dashboard[f'B{row}'].number_format = '#,##0.00'
    
    # 使用说明
    dashboard['E12'] = "使用说明"
    dashboard['E12'].font = Font(name='Arial', size=12, bold=True, color='2E75B6')
    
    instructions = [
        ("E13", "1. 在'流水账'表中记录每日收支"),
        ("E14", "2. 在'预算管理'表中设置预算"),
        ("E15", "3. 在'分类设置'中自定义分类"),
        ("E16", "4. 在'资产负债表'中记录资产"),
        ("E17", "5. 所有数据自动关联计算"),
        ("E18", "6. 使用数据验证确保准确性")
    ]
    
    for cell, text in instructions:
        dashboard[cell] = text
        dashboard[cell].font = Font(name='Arial', size=10)
    
    # ==================== 2. 流水账工作表 ====================
    print("  创建流水账...")
    ledger = wb.create_sheet("流水账")
    
    # 设置列宽
    column_widths = {'A': 8, 'B': 12, 'C': 15, 'D': 25, 'E': 10, 'F': 12, 'G': 15}
    for col, width in column_widths.items():
        ledger.column_dimensions[col].width = width
    
    # 标题行
    headers = ["序号", "日期", "分类", "项目", "类型", "金额", "支付方式"]
    for col, header in enumerate(headers, 1):
        cell = ledger.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.alignment = Alignment(horizontal='center')
    
    # 示例数据
    sample_data = [
        [1, "2026-03-01", "餐饮", "午餐", "支出", 45.00, "微信支付"],
        [2, "2026-03-01", "交通", "地铁卡充值", "支出", 100.00, "支付宝"],
        [3, "2026-03-02", "工资", "3月工资", "收入", 15000.00, "银行转账"],
        [4, "2026-03-02", "购物", "购买书籍", "支出", 89.00, "信用卡"],
        [5, "2026-03-03", "娱乐", "电影票", "支出", 120.00, "微信支付"],
        [6, "2026-03-04", "餐饮", "晚餐", "支出", 68.00, "支付宝"],
        [7, "2026-03-05", "其他收入", "兼职收入", "收入", 800.00, "现金"]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ledger.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # 设置格式
            if col_idx == 6:  # 金额列
                cell.number_format = '#,##0.00'
                if row_data[4] == "支出":
                    cell.font = Font(color='FF0000')  # 支出用红色
                else:
                    cell.font = Font(color='00B050')  # 收入用绿色
    
    # 添加公式行
    total_row = len(sample_data) + 2
    ledger.cell(row=total_row, column=5, value="总收入:")
    ledger.cell(row=total_row, column=5).font = Font(bold=True)
    ledger.cell(row=total_row, column=6, value='=SUMIF(F:F,"收入",G:G)')
    ledger.cell(row=total_row, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row, column=6).font = Font(bold=True, color='00B050')
    
    ledger.cell(row=total_row+1, column=5, value="总支出:")
    ledger.cell(row=total_row+1, column=5).font = Font(bold=True)
    ledger.cell(row=total_row+1, column=6, value='=SUMIF(F:F,"支出",G:G)')
    ledger.cell(row=total_row+1, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row+1, column=6).font = Font(bold=True, color='FF0000')
    
    ledger.cell(row=total_row+2, column=5, value="结余:")
    ledger.cell(row=total_row+2, column=5).font = Font(bold=True)
    ledger.cell(row=total_row+2, column=6, value=f'=G{total_row}-G{total_row+1}')
    ledger.cell(row=total_row+2, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row+2, column=6).font = Font(bold=True)
    
    # ==================== 3. 预算管理工作表 ====================
    print("  创建预算管理...")
    budget = wb.create_sheet("预算管理")
    
    # 设置列宽
    budget.column_dimensions['A'].width = 20
    budget.column_dimensions['B'].width = 12
    budget.column_dimensions['C'].width = 12
    budget.column_dimensions['D'].width = 12
    budget.column_dimensions['E'].width = 15
    budget.column_dimensions['F'].width = 15
    
    # 标题
    budget['A1'] = "月度预算管理"
    budget['A1'].font = Font(name='Arial', size=14, bold=True, color='2E75B6')
    budget.merge_cells('A1:F1')
    
    # 表头
    budget_headers = ["预算分类", "月度预算", "实际支出", "差额", "完成率", "状态"]
    for col, header in enumerate(budget_headers, 1):
        cell = budget.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='5B9BD5')
        cell.alignment = Alignment(horizontal='center')
    
    # 预算分类数据
    budget_categories = [
        ["餐饮", 1500],
        ["交通", 800],
        ["购物", 2000],
        ["娱乐", 500],
        ["住房", 3000],
        ["医疗", 300],
        ["教育", 1000],
        ["其他", 500]
    ]
    
    for row_idx, (category, amount) in enumerate(budget_categories, 4):
        # 分类
        budget.cell(row=row_idx, column=1, value=category)
        
        # 月度预算
        budget.cell(row=row_idx, column=2, value=amount)
        budget.cell(row=row_idx, column=2).number_format = '#,##0.00'
        
        # 实际支出（占位符）
        budget.cell(row=row_idx, column=3, value=0)
        budget.cell(row=row_idx, column=3).number_format = '#,##0.00'
        
        # 差额公式
        budget.cell(row=row_idx, column=4, value=f'=B{row_idx}-C{row_idx}')
        budget.cell(row=row_idx, column=4).number_format = '#,##0.00'
        
        # 完成率公式
        budget.cell(row=row_idx, column=5, value=f'=IF(B{row_idx}>0,C{row_idx}/B{row_idx},0)')
        budget.cell(row=row_idx, column=5).number_format = '0.0%'
        
        # 状态公式
        budget.cell(row=row_idx, column=6, value=f'=IF(D{row_idx}>=0,"在预算内",IF(D{row_idx}>=-B{row_idx}*0.1,"接近超支","已超支"))')
    
    # 总计行
    total_row = len(budget_categories) + 4
    budget.cell(row=total_row, column=1, value="总计")
    budget.cell(row=total_row, column=1).font = Font(bold=True)
    
    for col in [2, 3, 4]:
        budget.cell(row=total_row, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})')
        budget.cell(row=total_row, column=col).number_format = '#,##0.00'
        budget.cell(row=total_row, column=col).font = Font(bold=True)
    
    # ==================== 4. 分类设置工作表 ====================
    print("  创建分类设置...")
    categories = wb.create_sheet("分类设置")
    
    categories['A1'] = "收支分类设置"
    categories['A1'].font = Font(name='Arial', size=14, bold=True, color='2E75B6')
    categories.merge_cells('A1:C1')
    
    # 收入分类
    categories['A3'] = "收入分类"
    categories['A3'].font = Font(name='Arial', size=12, bold=True)
    
    income_cats = ["工资", "奖金", "投资收入", "兼职收入", "其他收入", "退款", "报销"]
    for i, cat in enumerate(income_cats, 4):
        categories.cell(row=i, column=1, value=cat)
    
    # 支出分类
    categories['B3'] = "支出分类"
    categories['B3'].font = Font(name='Arial', size=12, bold=True)
    
    expense_cats = ["餐饮", "交通", "购物", "娱乐", "住房", "医疗", "教育", "通讯", "保险", "其他"]
    for i, cat in enumerate(expense_cats, 4):
        categories.cell(row=i, column=2, value=cat)
    
    # 支付方式
    categories['C3'] = "支付方式"
    categories['C3'].font = Font(name='Arial', size=12, bold=True)
    
    payment_methods = ["现金", "微信支付", "支付宝", "信用卡", "银行转账", "其他"]
    for i, method in enumerate(payment_methods, 4):
        categories.cell(row=i, column=3, value=method)
    
    # ==================== 5. 资产负债表 ====================
    print("  创建资产负债表...")
    balance_sheet = wb.create_sheet("资产负债表")
    
    balance_sheet['A1'] = "个人资产负债表"
    balance_sheet['A1'].font = Font(name='Arial', size=14, bold=True, color='2E75B6')
    balance_sheet.merge_cells('A1:F1')
    
    # 资产表头
    balance_sheet['A3'] = "资产"
    balance_sheet['A3'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    balance_sheet['A3'].fill = PatternFill('solid', start_color='70AD47')
    balance_sheet.merge_cells('A3:C3')
    
    asset_headers = ["资产项目", "金额", "备注"]
    for col, header in enumerate(asset_headers, 1):
        cell = balance_sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='E2EFDA')
    
    # 负债表头
    balance_sheet['D3'] = "负债"
    balance_sheet['D3'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    balance_sheet['D3'].fill = PatternFill('solid', start_color='C00000')
    balance_sheet.merge_cells('D3:F3')
    
    liability_headers = ["负债项目", "金额", "备注"]
    for col, header in enumerate(liability_headers, 4):
        cell = balance_sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FCE4D6')
    
    # 示例资产数据
    assets = [
        ["现金", 5000, "日常备用金"],
        ["银行存款", 50000, "主要储蓄账户"],
        ["投资账户", 30000, "股票基金投资"],
        ["公积金", 80000, "住房公积金"],
        ["其他资产", 5000, "贵重物品"]
    ]
    
    for row_idx, (item, amount, note) in enumerate(assets, 5):
        balance_sheet.cell(row=row_idx, column=1, value=item)
        balance_sheet.cell(row=row_idx, column=2, value=amount)
        balance_sheet.cell(row=row_idx, column=2).number_format = '#,##0.00'
        balance_sheet.cell(row=row_idx, column=3, value=note)
    
    # 示例负债数据
    liabilities = [
        ["信用卡欠款", 5000, "本月待还"],
        ["房贷", 300000, "房屋贷款"],
        ["车贷", 80000, "汽车贷款"],
        ["其他负债", 2000, "个人借款"]
    ]
    
    for row_idx, (item, amount, note) in enumerate(liabilities, 5):
        balance_sheet.cell(row=row_idx, column=4, value=item)
        balance_sheet.cell(row=row_idx, column=5, value=amount)
        balance_sheet.cell(row=row_idx, column=5).number_format = '#,##0.00'
        balance_sheet.cell(row=row_idx, column=6, value=note)
    
    # 计算总计
    asset_total_row = len(assets) + 5
    liability_total_row = len(liabilities) + 5
    
    # 资产总计
    balance_sheet.cell(row=asset_total_row, column=1, value="资产总计")
    balance_sheet.cell(row=asset_total_row, column=1).font = Font(bold=True)
    balance_sheet.cell(row=asset_total_row, column=2, value=f'=SUM(B5:B{asset_total_row-1})')
    balance_sheet.cell(row=asset_total_row, column=2).number_format = '#,##0.00'
    balance_sheet.cell(row=asset_total_row, column=2).font = Font(bold=True)
    
    # 负债总计
    balance_sheet.cell(row=liability_total_row, column=4, value="负债总计")
    balance_sheet.cell(row=liability_total_row, column=4).font = Font(bold=True)
    balance_sheet.cell(row=liability_total_row, column=5, value=f'=SUM(E5:E{liability_total_row-1})')
    balance_sheet.cell(row=liability_total_row, column=5).number_format = '#,##0.00'
    balance_sheet.cell(row=liability_total_row, column=5).font = Font(bold=True)
    
    # 净资产
    net_worth_row = max(asset_total_row, liability_total_row) + 1
    balance_sheet.cell(row=net_worth_row, column=1, value="净资产")
    balance_sheet.cell(row=net_worth_row, column=1).font = Font(bold=True, size=12)
    balance_sheet.cell(row=net_worth_row, column=2, value=f'=B{asset_total_row}-E{liability_total_row}')
    balance_sheet.cell(row=net_worth_row, column=2).number_format = '#,##0.00'
    balance_sheet.cell(row=net_worth_row, column=2).font = Font(bold=True, size=12, color='00B050')
    
    # ==================== 6. 使用说明工作表 ====================
    print("  创建使用说明...")
    instructions = wb.create_sheet("使用说明")
    
    instructions['A1'] = "专业记账模板使用说明"
    instructions['A1'].font = Font(name='Arial', size=16, bold=True, color='2E75B6')
    instructions.merge_cells('A1:D1')
    
    sections = [
        ("📊 模板结构", [
            "1. 仪表板: 关键指标概览",
            "2. 流水账: 记录每日收支明细",
            "3. 预算管理: 设置和跟踪预算",
            "4. 分类设置: 自定义收支分类",
            "5. 资产负债表: 记录资产和负债"
        ]),
        ("📝 使用步骤", [
            "1. 在'分类设置'中自定义分类",
            "2. 在'流水账'中记录每日收支",
            "3. 在'预算管理'中设置预算",
            "4. 在'资产负债表'中更新资产",
            "5. 查看'仪表板'了解财务状况"
        ]),
        ("🔧 功能特点", [
            "• 自动计算收支结余",
            "• 预算执行情况跟踪",
            "• 多维度数据分析",
            "• 支持自定义分类",
            "• 专业财务报表格式"
        ]),
        ("💡 使用技巧", [
            "• 每日及时记录收支",
            "• 定期核对银行账单",
            "• 设置合理的预算目标",
            "• 利用分类分析消费习惯",
            "• 定期备份Excel文件"
        ]),
        ("🎯 专业功能", [
            "• 收支分类统计分析",
            "• 预算与实际对比",
            "• 资产净值计算",
            "• 月度年度汇总",
            "• 数据可视化准备"
        ])
    ]
    
    current_row = 3
    for title, items in sections:
        # 标题
        instructions.cell(row=current_row, column=1, value=title)
        instructions.cell(row=current_row, column=1).font = Font(name='Arial', size=12, bold=True, color='4472C4')
        current_row += 1
        
        # 内容
        for item in items:
            instructions.cell(row=current_row, column=1, value=item)
            instructions.cell(row=current_row, column=1).font = Font(name='Arial', size=10)
            current_row += 1
        
        current_row += 1  # 空行
    
    # 设置列宽
    instructions.column_dimensions['A'].width = 50
    
    # ==================== 7. 月度报表工作表 ====================
    print("  创建月度报表...")
    monthly_report = wb.create_sheet("月度报表")
    
    monthly_report['A1'] = "月度收支报表"
    monthly_report['A1'].font = Font(name='Arial', size=14, bold=True, color='2E75B6')
    monthly_report.merge_cells('A1:G1')
    
    # 表头
    report_headers = ["分类", "预算金额", "实际支出", "差额", "预算完成率", "趋势分析", "改进建议"]
    for col, header in enumerate(report_headers, 1):
        cell = monthly_report.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.alignment = Alignment(horizontal='center')
    
    # 设置列宽
    monthly_report.column_dimensions['A'].width = 15
    monthly_report.column_dimensions['B'].width = 12
    monthly_report.column_dimensions['C'].width = 12
    monthly_report.column_dimensions['D'].width = 12
    monthly_report.column_dimensions['E'].width = 15
    monthly_report.column_dimensions['F'].width = 15
    monthly_report.column_dimensions['G'].width = 25
    
    # 示例数据
    report_data = [
        ["餐饮", 1500, 1200, "=B4-C4", "=IF(B4>0,C4/B4,0)", "良好", "继续保持"],
        ["交通", 800, 750, "=B5-C5", "=IF(B5>0,C5/B5,0)", "良好", "合理控制"],
        ["购物", 2000, 2200, "=B6-C6", "=IF(B6>0,C6/B6,0)", "超支", "减少非必要购物"],
        ["娱乐", 500, 300, "=B7-C7", "=IF(B7>0,C7/B7,0)", "优秀", "继续保持"],
        ["住房", 3000, 3000, "=B8-C8", "=IF(B8>0,C8/B8,0)", "达标", "固定支出"],
        ["医疗", 300, 150, "=B9-C9", "=IF(B9>0,C9/B9,0)", "优秀", "健康管理良好"],
        ["教育", 1000, 800, "=B10-C10", "=IF(B10>0,C10/B10,0)", "良好", "合理投资"],
        ["其他", 500, 600, "=B11-C11", "=IF(B11>0,C11/B11,0)", "注意", "控制杂项支出"]
    ]
    
    for row_idx, row_data in enumerate(report_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = monthly_report.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # 设置格式
            if col_idx in [2, 3, 4]:  # 金额列
                cell.number_format = '#,##0.00'
            elif col_idx == 5:  # 百分比列
                cell.number_format = '0.0%'
    
    # ==================== 保存文件 ====================
    # 重新排序工作表
    sheet_order = ["仪表板", "流水账", "预算管理", "分类设置", "资产负债表", "月度报表", "使用说明"]
    
    # 按指定顺序重新排列工作表
    ordered_sheets = []
    for sheet_name in sheet_order:
        if sheet_name in wb.sheetnames:
            ordered_sheets.append(wb[sheet_name])
    
    # 添加其他工作表
    for sheet in wb.worksheets:
        if sheet.title not in sheet_order:
            ordered_sheets.append(sheet)
    
    wb._sheets = ordered_sheets
    
    # 保存文件
    filename = f"专业记账模板_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    
    # 获取文件大小
    file_size = os.path.getsize(filename)
    
    print(f"\n✅ 专业记账模板已创建: {filename}")
    print(f"📊 包含 {len(wb.worksheets)} 个工作表: {', '.join([ws.title for ws in wb.worksheets])}")
    print(f"📁 文件大小: {file_size // 1024} KB")
    print(f"🎨 专业功能:")
    print(f"   • 收支流水账记录")
    print(f"   • 预算管理与跟踪")
    print(f"   • 资产负债表")
    print(f"   • 分类统计分析")
    print(f"   • 月度报表分析")
    print(f"   • 数据可视化准备")
    
    return filename

if __name__ == "__main__":
    try:
        print("=" * 50)
        print("🎯 专业记账模板生成器")
        print("=" * 50)
        
        filename = create_simple_accounting_template()
        
        print("\n" + "=" * 50)
        print("🎉 模板创建成功！")
        print("=" * 50)
        print(f"\n📂 文件位置: {os.path.abspath(filename)}")
        print("\n💡 使用建议:")
        print("1. 打开Excel文件")
        print("2. 在'分类设置'中自定义分类")
        print("3. 在'流水账'中记录收支")
        print("4. 在'预算管理'中设置预算")
        print("5. 查看'仪表板'了解财务状况")
        print("\n🔧 专业功能已内置:")
        print("• 自动计算公式")
        print("• 预算执行跟踪")
        print("• 财务报表生成")
        print("• 数据统计分析")
        
    except Exception as e:
        print(f"\n❌ 创建模板时出错: {e}")
        import traceback
        traceback.print_exc()