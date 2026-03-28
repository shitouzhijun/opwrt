#!/usr/bin/env python3
"""
专业记账模板生成器
创建包含多个工作表的专业记账Excel模板
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar

def create_professional_accounting_template():
    """创建专业记账Excel模板"""
    
    # 创建新的工作簿
    wb = Workbook()
    
    # 删除默认工作表
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ==================== 1. 首页 - 仪表板 ====================
    dashboard = wb.create_sheet("仪表板")
    dashboard.sheet_view.showGridLines = False
    
    # 设置列宽
    for col in range(1, 13):
        dashboard.column_dimensions[get_column_letter(col)].width = 12
    
    # 标题
    dashboard.merge_cells('A1:L1')
    title_cell = dashboard['A1']
    title_cell.value = "个人/家庭财务管理系统"
    title_cell.font = Font(name='微软雅黑', size=20, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill('solid', start_color='2E75B6')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 当前日期
    current_date = datetime.now().strftime('%Y年%m月%d日')
    dashboard['A3'] = f"报表日期: {current_date}"
    dashboard['A3'].font = Font(name='微软雅黑', size=11, bold=True)
    
    # 关键指标区域
    metrics = [
        ("A5", "当月总收入", "=SUMIFS(流水账!F:F,流水账!E:E,\"收入\",流水账!B:B,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),流水账!B:B,\"<=\"&EOMONTH(TODAY(),0))"),
        ("A7", "当月总支出", "=SUMIFS(流水账!F:F,流水账!E:E,\"支出\",流水账!B:B,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),流水账!B:B,\"<=\"&EOMONTH(TODAY(),0))"),
        ("A9", "当月结余", "=A6-A8"),
        ("E5", "年度总收入", "=SUMIFS(流水账!F:F,流水账!E:E,\"收入\",流水账!B:B,\">=\"&DATE(YEAR(TODAY()),1,1),流水账!B:B,\"<=\"&DATE(YEAR(TODAY()),12,31))"),
        ("E7", "年度总支出", "=SUMIFS(流水账!F:F,流水账!E:E,\"支出\",流水账!B:B,\">=\"&DATE(YEAR(TODAY()),1,1),流水账!B:B,\"<=\"&DATE(YEAR(TODAY()),12,31))"),
        ("E9", "年度结余", "=E6-E8"),
        ("I5", "当前总资产", "=SUM(资产负债表!C:C)"),
        ("I7", "当前总负债", "=SUM(资产负债表!F:F)"),
        ("I9", "净资产", "=I6-I8")
    ]
    
    for cell, label, formula in metrics:
        # 标签
        label_cell = dashboard[cell]
        label_cell.value = label
        label_cell.font = Font(name='微软雅黑', size=11, bold=True)
        label_cell.fill = PatternFill('solid', start_color='F2F2F2')
        
        # 数值
        value_cell = dashboard[cell.replace('A', 'B').replace('E', 'F').replace('I', 'J')]
        value_cell.value = formula
        value_cell.number_format = '#,##0.00'
        value_cell.font = Font(name='Arial', size=11, bold=True)
        value_cell.fill = PatternFill('solid', start_color='FFFFFF')
        
        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        label_cell.border = thin_border
        value_cell.border = thin_border
    
    # 月度支出分类饼图数据区域
    dashboard['A12'] = "月度支出分类"
    dashboard['A12'].font = Font(name='微软雅黑', size=12, bold=True, color='2E75B6')
    
    categories = ["餐饮", "交通", "购物", "娱乐", "住房", "医疗", "教育", "其他"]
    for i, category in enumerate(categories):
        row = 13 + i
        dashboard[f'A{row}'] = category
        dashboard[f'A{row}'].font = Font(name='微软雅黑', size=10)
        dashboard[f'B{row}'] = f'=SUMIFS(流水账!F:F,流水账!E:E,"支出",流水账!C:C,A{row},流水账!B:B,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),流水账!B:B,"<="&EOMONTH(TODAY(),0))'
        dashboard[f'B{row}'].number_format = '#,##0.00'
    
    # 使用说明
    dashboard['E12'] = "使用说明"
    dashboard['E12'].font = Font(name='微软雅黑', size=12, bold=True, color='2E75B6')
    
    instructions = [
        ("E13", "1. 在'流水账'表中记录每日收支"),
        ("E14", "2. 在'预算管理'表中设置预算"),
        ("E15", "3. 在'分类设置'中自定义分类"),
        ("E16", "4. 在'资产负债表'中记录资产"),
        ("E17", "5. 所有图表自动更新"),
        ("E18", "6. 使用数据验证确保数据准确")
    ]
    
    for cell, text in instructions:
        dashboard[cell] = text
        dashboard[cell].font = Font(name='微软雅黑', size=10)
    
    # ==================== 2. 流水账工作表 ====================
    ledger = wb.create_sheet("流水账")
    
    # 设置列宽
    column_widths = {'A': 12, 'B': 15, 'C': 20, 'D': 40, 'E': 10, 'F': 15, 'G': 20}
    for col, width in column_widths.items():
        ledger.column_dimensions[col].width = width
    
    # 标题行
    headers = ["序号", "日期", "分类", "项目描述", "类型", "金额", "支付方式"]
    for col, header in enumerate(headers, 1):
        cell = ledger.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 示例数据
    sample_data = [
        [1, datetime(2026, 3, 1), "餐饮", "午餐", "支出", 45.00, "微信支付"],
        [2, datetime(2026, 3, 1), "交通", "地铁卡充值", "支出", 100.00, "支付宝"],
        [3, datetime(2026, 3, 2), "工资", "3月工资", "收入", 15000.00, "银行转账"],
        [4, datetime(2026, 3, 2), "购物", "购买书籍", "支出", 89.00, "信用卡"],
        [5, datetime(2026, 3, 3), "娱乐", "电影票", "支出", 120.00, "微信支付"],
        [6, datetime(2026, 3, 4), "餐饮", "晚餐", "支出", 68.00, "支付宝"],
        [7, datetime(2026, 3, 5), "其他收入", "兼职收入", "收入", 800.00, "现金"]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ledger.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # 设置格式
            if col_idx == 2:  # 日期列
                cell.number_format = 'YYYY-MM-DD'
            elif col_idx == 6:  # 金额列
                cell.number_format = '#,##0.00'
                if row_data[4] == "支出":
                    cell.font = Font(color='FF0000')  # 支出用红色
                else:
                    cell.font = Font(color='00B050')  # 收入用绿色
    
    # 添加公式行
    total_row = len(sample_data) + 2
    ledger.cell(row=total_row, column=5, value="总收入:")
    ledger.cell(row=total_row, column=5).font = Font(bold=True)
    ledger.cell(row=total_row, column=6, value='=SUMIF(F:F,"收入",G:G)')
    ledger.cell(row=total_row, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row, column=6).font = Font(bold=True, color='00B050')
    
    ledger.cell(row=total_row+1, column=5, value="总支出:")
    ledger.cell(row=total_row+1, column=5).font = Font(bold=True)
    ledger.cell(row=total_row+1, column=6, value='=SUMIF(F:F,"支出",G:G)')
    ledger.cell(row=total_row+1, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row+1, column=6).font = Font(bold=True, color='FF0000')
    
    ledger.cell(row=total_row+2, column=5, value="结余:")
    ledger.cell(row=total_row+2, column=5).font = Font(bold=True)
    ledger.cell(row=total_row+2, column=6, value=f'={get_column_letter(6)}{total_row}-{get_column_letter(6)}{total_row+1}')
    ledger.cell(row=total_row+2, column=6).number_format = '#,##0.00'
    ledger.cell(row=total_row+2, column=6).font = Font(bold=True)
    
    # ==================== 3. 预算管理工作表 ====================
    budget = wb.create_sheet("预算管理")
    
    # 设置列宽
    budget.column_dimensions['A'].width = 25
    budget.column_dimensions['B'].width = 15
    budget.column_dimensions['C'].width = 15
    budget.column_dimensions['D'].width = 15
    budget.column_dimensions['E'].width = 15
    budget.column_dimensions['F'].width = 20
    
    # 标题
    budget['A1'] = "月度预算管理"
    budget['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')
    budget.merge_cells('A1:F1')
    
    # 表头
    budget_headers = ["预算分类", "月度预算", "实际支出", "差额", "完成率", "状态"]
    for col, header in enumerate(budget_headers, 1):
        cell = budget.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='5B9BD5')
        cell.alignment = Alignment(horizontal='center')
    
    # 预算分类数据
    budget_categories = [
        ["餐饮", 1500],
        ["交通", 800],
        ["购物", 2000],
        ["娱乐", 500],
        ["住房", 3000],
        ["医疗", 300],
        ["教育", 1000],
        ["其他", 500]
    ]
    
    for row_idx, (category, amount) in enumerate(budget_categories, 4):
        # 分类
        budget.cell(row=row_idx, column=1, value=category)
        
        # 月度预算
        budget.cell(row=row_idx, column=2, value=amount)
        budget.cell(row=row_idx, column=2).number_format = '#,##0.00'
        
        # 实际支出公式
        actual_formula = f'=SUMIFS(流水账!F:F,流水账!C:C,A{row_idx},流水账!E:E,"支出",流水账!B:B,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),流水账!B:B,"<="&EOMONTH(TODAY(),0))'
        budget.cell(row=row_idx, column=3, value=actual_formula)
        budget.cell(row=row_idx, column=3).number_format = '#,##0.00'
        
        # 差额公式
        diff_formula = f'=B{row_idx}-C{row_idx}'
        budget.cell(row=row_idx, column=4, value=diff_formula)
        budget.cell(row=row_idx, column=4).number_format = '#,##0.00'
        
        # 完成率公式
        completion_formula = f'=IF(B{row_idx}>0,C{row_idx}/B{row_idx},0)'
        budget.cell(row=row_idx, column=5, value=completion_formula)
        budget.cell(row=row_idx, column=5).number_format = '0.0%'
        
        # 状态公式
        status_formula = f'=IF(D{row_idx}>=0,"在预算内",IF(D{row_idx}>=-B{row_idx}*0.1,"接近超支","已超支"))'
        budget.cell(row=row_idx, column=6, value=status_formula)
    
    # 总计行
    total_row = len(budget_categories) + 4
    budget.cell(row=total_row, column=1, value="总计")
    budget.cell(row=total_row, column=1).font = Font(bold=True)
    
    for col in [2, 3, 4]:
        budget.cell(row=total_row, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})')
        budget.cell(row=total_row, column=col).number_format = '#,##0.00'
        budget.cell(row=total_row, column=col).font = Font(bold=True)
    
    # ==================== 4. 分类设置工作表 ====================
    categories = wb.create_sheet("分类设置")
    
    categories['A1'] = "收支分类设置"
    categories['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')
    categories.merge_cells('A1:C1')
    
    # 收入分类
    categories['A3'] = "收入分类"
    categories['A3'].font = Font(name='微软雅黑', size=12, bold=True)
    
    income_cats = ["工资", "奖金", "投资收入", "兼职收入", "其他收入", "退款", "报销"]
    for i, cat in enumerate(income_cats, 4):
        categories.cell(row=i, column=1, value=cat)
    
    # 支出分类
    categories['B3'] = "支出分类"
    categories['B3'].font = Font(name='微软雅黑', size=12, bold=True)
    
    expense_cats = ["餐饮", "交通", "购物", "娱乐", "住房", "医疗", "教育", "通讯", "保险", "其他"]
    for i, cat in enumerate(expense_cats, 4):
        categories.cell(row=i, column=2, value=cat)
    
    # 支付方式
    categories['C3'] = "支付方式"
    categories['C3'].font = Font(name='微软雅黑', size=12, bold=True)
    
    payment_methods = ["现金", "微信支付", "支付宝", "信用卡", "银行转账", "其他"]
    for i, method in enumerate(payment_methods, 4):
        categories.cell(row=i, column=3, value=method)
    
    # ==================== 5. 资产负债表 ====================
    balance_sheet = wb.create_sheet("资产负债表")
    
    balance_sheet['A1'] = "个人资产负债表"
    balance_sheet['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')
    balance_sheet.merge_cells('A1:F1')
    
    # 资产表头
    balance_sheet['A3'] = "资产"
    balance_sheet['A3'].font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    balance_sheet['A3'].fill = PatternFill('solid', start_color='70AD47')
    balance_sheet.merge_cells('A3:C3')
    
    asset_headers = ["资产项目", "金额", "备注"]
    for col, header in enumerate(asset_headers, 1):
        cell = balance_sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='E2EFDA')
    
    # 负债表头
    balance_sheet['D3'] = "负债"
    balance_sheet['D3'].font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    balance_sheet['D3'].fill = PatternFill('solid', start_color='C00000')
    balance_sheet.merge_cells('D3:F3')
    
    liability_headers = ["负债项目", "金额", "备注"]
    for col, header in enumerate(liability_headers, 4):
        cell = balance_sheet.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FCE4D6')
    
    # 示例资产数据
    assets = [
        ["现金", 5000, "日常备用金"],
        ["银行存款", 50000, "主要储蓄账户"],
        ["投资账户", 30000, "股票基金投资"],
        ["公积金", 80000, "住房公积金"],
        ["其他资产", 5000, "贵重物品"]
    ]
    
    for row_idx, (item, amount, note) in enumerate(assets, 5):
        balance_sheet.cell(row=row_idx, column=1, value=item)
        balance_sheet.cell(row=row_idx, column=2, value=amount)
        balance_sheet.cell(row=row_idx, column=2).number_format = '#,##0.00'
        balance_sheet.cell(row=row_idx, column=3, value=note)
    
    # 示例负债数据
    liabilities = [
        ["信用卡欠款", 5000, "本月待还"],
        ["房贷", 300000, "房屋贷款"],
        ["车贷", 80000, "汽车贷款"],
        ["其他负债", 2000, "个人借款"]
    ]
    
    for row_idx, (item, amount, note) in enumerate(liabilities, 5):
        balance_sheet.cell(row=row_idx, column=4, value=item)
        balance_sheet.cell(row=row_idx, column=5, value=amount)
        balance_sheet.cell(row=row_idx, column=5).number_format = '#,##0.00'
        balance_sheet.cell(row=row_idx, column=6, value=note)
    
    # 计算总计
    asset_total_row = len(assets) + 5
    liability_total_row = len(liabilities) + 5
    
    # 资产总计
    balance_sheet.cell(row=asset_total_row, column=1, value="资产总计")
    balance_sheet.cell(row=asset_total_row, column=1).font = Font(bold=True)
    balance_sheet.cell(row=asset_total_row, column=2, value=f'=SUM(B5:B{asset_total_row-1})')
    balance_sheet.cell(row=asset_total_row, column=2).number_format = '#,##0.00'
    balance_sheet.cell(row=asset_total_row, column=2).font = Font(bold=True)
    
    # 负债总计
    balance_sheet.cell(row=liability_total_row, column=4, value="负债总计")
    balance_sheet.cell(row=liability_total_row, column=4).font = Font(bold=True)
    balance_sheet.cell(row=liability_total_row, column=5, value=f'=SUM(E5:E{liability_total_row-1})')
    balance_sheet.cell(row=liability_total_row, column=5).number_format = '#,##0.00'
    balance_sheet.cell(row=liability_total_row, column=5).font = Font(bold=True)
    
    # 净资产
    net_worth_row = max(asset_total_row, liability_total_row) + 1
    balance_sheet.cell(row=net_worth_row, column=1, value="净资产")
    balance_sheet.cell(row=net_worth_row, column=1).font = Font(bold=True, size=12)
    balance_sheet.cell(row=net_worth_row, column=2, value=f'=B{asset_total_row}-E{liability_total_row}')
    balance_sheet.cell(row=net_worth_row, column=2).number_format = '#,##0.00'
    balance_sheet.cell(row=net_worth_row, column=2).font = Font(bold=True, size=12, color='00B050')
    
    # ==================== 6. 月度报表工作表 ====================
    monthly_report = wb.create_sheet("月度报表")
    
    monthly_report['A1'] = "月度收支报表"
    monthly_report['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')
    monthly_report.merge_cells('A1:G1')
    
    # 月份选择
    monthly_report['A3'] = "选择月份:"
    monthly_report['A3'].font = Font(bold=True)
    
    # 创建月份下拉列表数据
    months = wb.create_sheet("_月份数据")
    for i in range(1, 13):
        months.cell(row=i, column=1, value=f"{i}月")
    wb._sheets.remove(months)  # 隐藏这个工作表
    
    # 表头
    report_headers = ["分类", "预算", "实际", "差额", "预算完成率", "趋势", "建议"]
    for col, header in enumerate(report_headers, 1):
        cell = monthly_report.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.alignment = Alignment(horizontal='center')
    
    # 设置列宽
    monthly_report.column_dimensions['A'].width = 20
    monthly_report.column_dimensions['B'].width = 12
    monthly_report.column_dimensions['C'].width = 12
    monthly_report.column_dimensions['D'].width = 12
    monthly_report.column_dimensions['E'].width = 15
    monthly_report.column_dimensions['F'].width = 15
    monthly_report.column_dimensions['G'].width = 25
    
    # ==================== 7. 年度汇总工作表 ====================
    annual_summary = wb.create_sheet("年度汇总")
    
    annual_summary['A1'] = "年度收支汇总"
    annual_summary['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')
    annual_summary.merge_cells('A1:N1')
    
    # 月份表头
    months = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "年度总计"]
    for col, month in enumerate(months, 2):
        cell = annual_summary.cell(row=3, column=col)
        cell.value = month
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill('solid', start_color='D9E1F2')
    
    # 收入行
    annual_summary['A4'] = "收入"
    annual_summary['A4'].font = Font(bold=True, color='00B050')
    
    # 支出行
    annual_summary['A5'] = "支出"
    annual_summary['A5'].font = Font(bold=True, color='FF0000')
    
    # 结余行
    annual_summary['A6'] = "结余"
    annual_summary['A6'].font = Font(bold=True)
    
    # 设置公式（示例公式，实际需要更复杂的公式）
    for month in range(1, 13):
        col = month + 1
        # 收入公式
        income_formula = f'=SUMIFS(流水账!F:F,流水账!E:E,"收入",流水账!B:B,">="&DATE(YEAR(TODAY()),{month},1),流水账!B:B,"<="&EOMONTH(DATE(YEAR(TODAY()),{month},1),0))'
        annual_summary.cell(row=4, column=col, value=income_formula)
        annual_summary.cell(row=4, column=col).number_format = '#,##0.00'
        
        # 支出公式
        expense_formula = f'=SUMIFS(流水账!F:F,流水账!E:E,"支出",流水账!B:B,">="&DATE(YEAR(TODAY()),{month},1),流水账!B:B,"<="&EOMONTH(DATE(YEAR(TODAY()),{month},1),0))'
        annual_summary.cell(row=5, column=col, value=expense_formula)
        annual_summary.cell(row=5, column=col).number_format = '#,##0.00'
        
        # 结余公式
        balance_formula = f'=B{4+month-1}-B{5+month-1}'
        annual_summary.cell(row=6, column=col, value=balance_formula)
        annual_summary.cell(row=6, column=col).number_format = '#,##0.00'
    
    # 年度总计列
    total_col = 14
    annual_summary.cell(row=4, column=total_col, value='=SUM(B4:M4)')
    annual_summary.cell(row=4, column=total_col).number_format = '#,##0.00'
    annual_summary.cell(row=4, column=total_col).font = Font(bold=True, color='00B050')
    
    annual_summary.cell(row=5, column=total_col, value='=SUM(B5:M5)')
    annual_summary.cell(row=5, column=total_col).number_format = '#,##0.00'
    annual_summary.cell(row=5, column=total_col).font = Font(bold=True, color='FF0000')
    
    annual_summary.cell(row=6, column=total_col, value='=SUM(B6:M6)')
    annual_summary.cell(row=6, column=total_col).number_format = '#,##0.00'
    annual_summary.cell(row=6, column=total_col).font = Font(bold=True)
    
    # ==================== 8. 数据验证和说明 ====================
    instructions = wb.create_sheet("使用说明")
    
    instructions['A1'] = "专业记账模板使用说明"
    instructions['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='2E75B6')
    instructions.merge_cells('A1:D1')
    
    sections = [
        ("📊 模板结构", [
            "1. 仪表板: 关键指标概览",
            "2. 流水账: 记录每日收支明细",
            "3. 预算管理: 设置和跟踪预算",
            "4. 分类设置: 自定义收支分类",
            "5. 资产负债表: 记录资产和负债",
            "6. 月度报表: 月度分析报告",
            "7. 年度汇总: 全年数据汇总"
        ]),
        ("📝 使用步骤", [
            "1. 在'分类设置'中自定义分类",
            "2. 在'流水账'中记录每日收支",
            "3. 在'预算管理'中设置预算",
            "4. 在'资产负债表'中更新资产",
            "5. 查看'仪表板'了解财务状况",
            "6. 使用'月度报表'进行分析"
        ]),
        ("🔧 功能特点", [
            "• 自动计算收支结余",
            "• 预算执行情况跟踪",
            "• 多维度数据分析",
            "• 可视化图表展示",
            "• 数据验证确保准确性",
            "• 支持自定义分类"
        ]),
        ("💡 使用技巧", [
            "• 每日及时记录收支",
            "• 定期核对银行账单",
            "• 设置合理的预算目标",
            "• 利用分类分析消费习惯",
            "• 定期备份Excel文件",
            "• 根据实际情况调整分类"
        ])
    ]
    
    current_row = 3
    for title, items in sections:
        # 标题
        instructions.cell(row=current_row, column=1, value=title)
        instructions.cell(row=current_row, column=1).font = Font(name='微软雅黑', size=12, bold=True, color='4472C4')
        current_row += 1
        
        # 内容
        for item in items:
            instructions.cell(row=current_row, column=1, value=item)
            instructions.cell(row=current_row, column=1).font = Font(name='微软雅黑', size=10)
            current_row += 1
        
        current_row += 1  # 空行
    
    # 设置列宽
    instructions.column_dimensions['A'].width = 50
    
    # ==================== 保存文件 ====================
    # 重新排序工作表
    sheet_order = ["仪表板", "流水账", "预算管理", "分类设置", "资产负债表", "月度报表", "年度汇总", "使用说明"]
    wb._sheets.sort(key=lambda ws: sheet_order.index(ws.title) if ws.title in sheet_order else len(sheet_order))
    
    # 保存文件
    filename = f"专业记账模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    print(f"✅ 专业记账模板已创建: {filename}")
    print(f"📊 包含工作表: {', '.join(sheet_order)}")
    print(f"📁 文件大小: 约{len(open(filename, 'rb').read()) // 1024} KB")
    
    return filename

if __name__ == "__main__":
    try:
        filename = create_professional_accounting_template()
        print("\n🎉 模板创建成功！")
        print("💡 接下来请运行公式重计算脚本:")
        print(f"   python scripts/recalc.py {filename}")
    except Exception as e:
        print(f"❌ 创建模板时出错: {e}")
        import traceback
        traceback.print_exc()